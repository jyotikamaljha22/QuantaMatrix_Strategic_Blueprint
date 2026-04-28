import base64
import csv
import hashlib
import math
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# =========================
# Page config
# =========================
st.set_page_config(
    page_title="Strategic Market Research | Ammonia Storage Tank",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================
# Theme Colors & Notes
# =========================
BURGUNDY = "#5B0F2E"
BURGUNDY_DARK = "#431022"
BURGUNDY_MID = "#7A1C41"
BURGUNDY_SOFT = "#A45A7B"
GOLD = "#C9A227"
ROSE = "#EEDBE4"
BG = "#F8F5F6"
CARD = "rgba(255, 255, 255, 0.85)"
INK = "#1A1014"
MUTED = "#6B5B63"
BORDER = "rgba(255, 255, 255, 0.9)"

PREVIEW_NOTE = (
    "This dashboard is for preview purposes only. Data shown here is illustrative, "
    "aggregated, and intentionally limited."
)

# =========================
# CSS / UI polish (Ultra-Premium Glassmorphism)
# =========================
st.markdown(
    f"""
    <style>
    /* Premium Web Font */
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {{
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        color: {INK} !important;
    }}

    /* UI HIDING FOR WHITE-LABELING */
    [data-testid="stToolbar"] {{ display: none !important; }}
    .viewerBadge_container, #viewerBadge_container {{ display: none !important; }}
    footer {{ display: none !important; }}
    header[data-testid="stHeader"] {{ background: transparent !important; box-shadow: none !important; }}

    /* CUSTOM ELEGANT SCROLLBAR */
    ::-webkit-scrollbar {{
        width: 6px;
        height: 6px;
    }}
    ::-webkit-scrollbar-track {{
        background: transparent;
    }}
    ::-webkit-scrollbar-thumb {{
        background: rgba(91,15,46,0.15);
        border-radius: 10px;
    }}
    ::-webkit-scrollbar-thumb:hover {{
        background: rgba(91,15,46,0.3);
    }}

    /* LIVING GRADIENT BACKGROUND */
    .stApp {{
      background: radial-gradient(circle at 15% 0%, rgba(201,162,39,0.04) 0%, transparent 40%),
                  radial-gradient(circle at 85% 100%, rgba(91,15,46,0.03) 0%, transparent 40%),
                  linear-gradient(180deg, #FCFAFB 0%, #F4ECEF 100%);
      background-attachment: fixed;
    }}

    /* SIDEBAR REFINEMENT */
    [data-testid="stSidebar"] {{
        background: rgba(255,255,255,0.6) !important;
        backdrop-filter: blur(20px) !important;
        -webkit-backdrop-filter: blur(20px) !important;
        border-right: 1px solid rgba(91,15,46,0.08) !important;
    }}

    /* SLEEK BUTTON STYLING */
    .stButton > button {{
        background: linear-gradient(135deg, {BURGUNDY} 0%, {BURGUNDY_MID} 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 10px 24px !important;
        font-weight: 700 !important;
        font-size: 0.9rem !important;
        letter-spacing: 0.02em;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
        box-shadow: 0 4px 14px rgba(91,15,46,0.2) !important;
    }}
    .stButton > button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(91,15,46,0.3) !important;
    }}
    
    .stDownloadButton > button {{
        background: transparent !important;
        color: {BURGUNDY} !important;
        border: 1px solid rgba(91,15,46,0.2) !important;
        box-shadow: none !important;
        padding: 6px 16px !important;
        font-size: 0.85rem !important;
    }}
    .stDownloadButton > button:hover {{
        background: rgba(91,15,46,0.04) !important;
        border-color: {BURGUNDY} !important;
    }}

    /* INPUT FIELDS */
    .stTextInput input {{
        border-radius: 8px !important;
        border: 1px solid rgba(91,15,46,0.15) !important;
        padding: 12px 14px !important;
        background: rgba(255,255,255,0.8) !important;
        transition: all 0.2s ease !important;
    }}
    .stTextInput input:focus {{
        border-color: {GOLD} !important;
        background: #fff !important;
        box-shadow: 0 0 0 3px rgba(201,162,39,0.15) !important;
    }}

    /* APP-LIKE NAVIGATION */
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] {{
        padding: 10px 14px !important;
        margin-bottom: 6px !important;
        border-radius: 8px !important;
        transition: all 0.2s ease !important;
        cursor: pointer !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:hover {{
        background: rgba(91,15,46,0.04) !important;
        transform: translateX(3px) !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"]:has(input[checked]) {{
         background: #ffffff !important;
         border-left: 4px solid {BURGUNDY} !important;
         box-shadow: 0 4px 12px rgba(0,0,0,0.03) !important;
    }}
    [data-testid="stSidebar"] [data-testid="stRadio"] label[data-baseweb="radio"] p {{
        font-weight: 600 !important;
        font-size: 0.92rem !important;
        color: {INK} !important;
    }}

    /* FROSTED BRANDING HEADER */
    .smr-brand {{
      background: rgba(255,255,255,0.7);
      border: 1px solid {BORDER};
      border-radius: 16px;
      padding: 20px;
      box-shadow: 0 8px 24px rgba(0,0,0,0.03);
      margin-bottom: 24px;
      backdrop-filter: blur(16px);
      -webkit-backdrop-filter: blur(16px);
    }}
    
    .hero {{
      background: linear-gradient(135deg, {BURGUNDY} 0%, {BURGUNDY_MID} 60%, {BURGUNDY_SOFT} 100%);
      color: white;
      border: 1px solid rgba(255,255,255,0.1);
      border-radius: 20px;
      padding: 34px 40px;
      box-shadow: 0 20px 40px rgba(61,16,33,0.15), inset 0 1px 0 rgba(255,255,255,0.2);
      margin-bottom: 24px;
      animation: floatIn 0.6s cubic-bezier(0.25, 0.8, 0.25, 1);
      position: relative;
      overflow: hidden;
    }}
    .hero::before {{
      content: ''; position: absolute; top: -50%; right: -10%; width: 60%; height: 200%;
      background: radial-gradient(circle, rgba(255,255,255,0.08) 0%, transparent 60%);
      transform: rotate(-45deg);
    }}
    .hero-head {{
      display:flex; align-items:center; gap:18px; margin-bottom:8px; position: relative; z-index: 2;
    }}
    .hero h2 {{
      margin: 0; font-size: 2.2rem; font-weight: 800; letter-spacing: -0.02em; text-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }}
    
    /* GLASSMORPHISM DATA CARDS */
    .metric-card {{
      background: {CARD};
      backdrop-filter: blur(20px);
      -webkit-backdrop-filter: blur(20px);
      border: 1px solid {BORDER};
      border-radius: 16px;
      padding: 24px;
      box-shadow: 0 12px 36px rgba(91,15,46,0.03), 0 2px 8px rgba(0,0,0,0.02);
      min-height: 135px;
      animation: floatIn 0.6s cubic-bezier(0.25, 0.8, 0.25, 1);
      transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
    }}
    .metric-card:hover {{
      transform: translateY(-4px);
      box-shadow: 0 16px 40px rgba(91,15,46,0.08), 0 4px 12px rgba(0,0,0,0.03);
      border-color: rgba(201,162,39,0.3);
    }}
    .metric-label {{
      color: {MUTED};
      font-size: 0.82rem;
      margin-bottom: 8px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }}
    .metric-value {{
      color: {BURGUNDY};
      font-size: 2.6rem;
      font-weight: 800;
      line-height: 1;
      letter-spacing: -0.02em;
      margin-bottom: 12px;
    }}
    .metric-foot {{
      color: {MUTED};
      font-size: 0.85rem;
      line-height: 1.4;
    }}

    .section-card {{
      background: {CARD};
      backdrop-filter: blur(20px);
      -webkit-backdrop-filter: blur(20px);
      border: 1px solid {BORDER};
      border-radius: 20px;
      padding: 24px;
      box-shadow: 0 12px 36px rgba(91,15,46,0.03);
      margin-bottom: 24px;
      animation: floatIn 0.6s cubic-bezier(0.25, 0.8, 0.25, 1);
    }}
    .section-title {{
      display: flex; justify-content: space-between; align-items: center;
      font-size: 1.2rem; font-weight: 800; color: {BURGUNDY}; margin-bottom: 6px; letter-spacing: -0.01em;
    }}
    .section-sub {{
      color: {MUTED}; font-size: 0.9rem; margin-bottom: 20px; line-height: 1.5;
    }}
    
    .insight-box {{
      background: #ffffff;
      border: 1px solid rgba(201,162,39,0.2);
      border-left: 6px solid {GOLD};
      border-radius: 12px;
      padding: 20px 24px;
      color: {INK};
      font-size: 0.95rem;
      line-height: 1.6;
      margin-bottom: 24px;
      box-shadow: 0 8px 24px rgba(201,162,39,0.05);
      animation: floatIn 0.6s ease-out;
    }}
    
    .viewer-chip {{
      display:inline-block; padding:6px 12px; border-radius:999px; font-size:0.75rem; font-weight:700;
      background: rgba(91,15,46,0.06); border: 1px solid rgba(91,15,46,0.1); color: {BURGUNDY};
      margin-top:10px; margin-bottom: 10px;
    }}

    @keyframes floatIn {{
      from {{ opacity: 0; transform: translateY(16px); }}
      to {{ opacity: 1; transform: translateY(0px); }}
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

# =========================
# Core Data Helpers
# =========================
def normalize_text(value: str) -> str:
    return str(value).strip().replace("–", "-").replace("—", "-").replace("’", "'").replace("“", '"').replace("”", '"')

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [normalize_text(c) for c in out.columns]
    return out

def get_col(df: pd.DataFrame, possible_names: Iterable[str]) -> str | None:
    cols_norm = {normalize_text(c): c for c in df.columns}
    for name in possible_names:
        key = normalize_text(name)
        if key in cols_norm: return cols_norm[key]
    for name in possible_names:
        key = normalize_text(name)
        for k,v in cols_norm.items():
            if key in k: return v
    return df.columns[0] if len(df.columns)>0 else None

def find_row(ws, title: str):
    title_norm = normalize_text(title)
    for i in range(1, ws.max_row + 1):
        value = ws.cell(i, 1).value
        if value is not None and normalize_text(value) == title_norm: return i
    return None

def read_table(ws, title: str) -> pd.DataFrame:
    row = find_row(ws, title)
    if row is None: raise KeyError(f"Table not found: {title} in {ws.title}")
    header_row = row + 1
    headers = []
    c = 1
    while c <= ws.max_column and ws.cell(header_row, c).value is not None:
        headers.append(normalize_text(ws.cell(header_row, c).value))
        c += 1
    records = []
    r = header_row + 1
    while r <= ws.max_row and ws.cell(r, 1).value is not None:
        row_vals = [ws.cell(r, col).value for col in range(1, len(headers) + 1)]
        records.append(row_vals)
        r += 1
    df = pd.DataFrame(records, columns=headers)
    return normalize_columns(df)

# =========================
# UI Component Helpers
# =========================
def fmt_mn(value: float, mult: float = 1.0) -> str:
    if value is None or pd.isna(value): return "—"
    adj_val = value * mult
    if abs(adj_val) >= 1000: return f"≈${adj_val/1000:.1f}B"
    return f"≈${adj_val:,.0f}M"

def fmt_units(value: float, mult: float = 1.0) -> str:
    if value is None or pd.isna(value): return "—"
    return f"≈{value * mult:,.1f}"

def card_metric(label: str, value: str, foot: str):
    st.markdown(f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div><div class="metric-foot">{foot}</div></div>', unsafe_allow_html=True)

def section_open(title: str, subtitle: str = ""):
    st.markdown(
        f'<div class="section-card"><div class="section-title">{title}</div><div class="section-sub">{subtitle}</div>',
        unsafe_allow_html=True,
    )

def section_close():
    st.markdown("</div>", unsafe_allow_html=True)

def chart_theme(fig):
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=INK, family="'Plus Jakarta Sans', sans-serif"),
        margin=dict(l=10, r=10, t=40, b=10), legend_title_text="",
        legend=dict(orientation="h", yanchor="bottom", y=1.05, x=0),
        hoverlabel=dict(bgcolor="white", font_size=13, font_family="'Plus Jakarta Sans', sans-serif", bordercolor=BORDER),
    )
    fig.update_xaxes(showgrid=False, linecolor="rgba(0,0,0,0.05)", tickfont=dict(size=11, color=MUTED))
    fig.update_yaxes(gridcolor="rgba(91,15,46,0.04)", zeroline=False, tickfont=dict(size=11, color=MUTED))
    return fig

def load_logo_base64() -> str | None:
    candidates = [Path("smrlogonew.svg"), Path(__file__).with_name("smrlogonew.svg"), Path.cwd() / "smrlogonew.svg", Path("logo.svg"), Path(__file__).with_name("logo.svg")]
    for path in candidates:
        if path.exists():
            try: return base64.b64encode(path.read_bytes()).decode("utf-8")
            except Exception: pass
    return None

LOGO_B64 = load_logo_base64()

def brand_sidebar():
    logo_html = f'<img src="data:image/svg+xml;base64,{LOGO_B64}" style="height:48px; width:auto; margin-bottom:12px;" />' if LOGO_B64 else ""
    st.sidebar.markdown(f'<div class="smr-brand">{logo_html}<h1 style="color:{BURGUNDY}; margin:0; font-size:1.1rem; font-weight:800; letter-spacing:-0.02em; line-height:1.2;">Strategic Market Research</h1><p style="color:{MUTED}; margin:4px 0 0 0; font-size:0.82rem; font-weight:500;">Ammonia Storage Tank<br>Preview Dashboard</p></div>', unsafe_allow_html=True)

def page_header(title: str, subtitle: str):
    logo_html = f'<img src="data:image/svg+xml;base64,{LOGO_B64}" style="height:42px; width:auto; filter: brightness(0) invert(1);" />' if LOGO_B64 else ""
    st.markdown(f'<div class="hero"><div class="kicker" style="color:var(--gold); font-size:0.75rem; font-weight:800; letter-spacing:0.1em; text-transform:uppercase; margin-bottom:8px; display:inline-block;">Strategic Market Research</div><div class="hero-head">{logo_html}<h2>{title}</h2></div><p style="font-size: 1rem; color: rgba(255,255,255,0.9); margin:0; max-width:900px; line-height:1.5;">{subtitle}</p></div>', unsafe_allow_html=True)

def page_footer():
    st.markdown(f'<div style="margin-top:30px; text-align:center; padding: 24px; color:{MUTED}; font-size:0.8rem; border-top: 1px solid rgba(0,0,0,0.06);"><strong>Strategic Market Research</strong> &copy; {datetime.now().year} — {PREVIEW_NOTE}</div>', unsafe_allow_html=True)

def log_access(name: str, email: str):
    log_path = Path("smr_preview_access_log.csv")
    fieldnames = ["timestamp_utc", "viewer_name", "viewer_email"]
    record = {"timestamp_utc": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"), "viewer_name": name.strip(), "viewer_email": email.strip()}
    try:
        write_header = not log_path.exists()
        with log_path.open("a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if write_header: writer.writeheader()
            writer.writerow(record)
    except Exception: pass

# =========================
# Data Loading
# =========================
@st.cache_data(show_spinner=False)
def load_model(workbook_path: str):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    data = {}
    ws = wb["01_Control_Panel"]
    data["market_revenue"] = read_table(ws, "Global Ammonia Storage Tank Market Size (2025-2035, $Mn)")
    data["tank_demand"] = read_table(ws, "Global Ammonia Storage Tank Demand (2025-2035, Units)")
    data["storage_capacity"] = read_table(ws, "Global Storage Capacity Demand (2025-2035, '000 m³)")
    data["growth_summary"] = read_table(ws, "Market growth summary (2025-2035, %)")
    data["end_share"] = read_table(ws, "Market split by end market (2025 vs 2035, % of total revenue)")
    data["region_share"] = read_table(ws, "Market split by region (2025 vs 2035, % of total revenue)")
    data["scenario_compare"] = read_table(ws, "Scenario comparison - 2035 output view (Units & $Mn)")
    data["tk_addr_cp"] = read_table(ws, "thyssenkrupp addressable market (2025-2035, $Mn)")

    ws = wb["02_Global_Demand_Engine"]
    data["end_units"] = read_table(ws, "Tank demand by end market (2025 / 2030 / 2035, Units)")

    ws = wb["03_Project_Pipeline"]
    data["project_db"] = read_table(ws, "Named ammonia project pipeline database (selected rows; Units = expected tanks)")
    data["pipeline_region"] = read_table(ws, "Named pipeline by region (risk-adjusted capacity, Mtpa)")
    data["pipeline_end"] = read_table(ws, "Named pipeline by end market (risk-adjusted capacity, Mtpa)")
    data["stage_probs"] = read_table(ws, "Stage probability factors used for risk adjustment")
    data["top_projects"] = read_table(ws, "Top 10 named projects by expected tank demand (Units)")

    ws = wb["05_Pricing_Market_Size"]
    data["avg_tank_cost"] = read_table(ws, "Average tank cost by end market (2025 vs 2035, $Mn per tank)")
    data["pricing_index"] = read_table(ws, "Inflation / pricing index (2025-2035, 2025 = 100)")
    data["revenue_end"] = read_table(ws, "Revenue by end market (2025 / 2030 / 2035, $Mn)")
    data["revenue_comp"] = read_table(ws, "Revenue composition (2025 vs 2035, % of total market revenue)")

    ws = wb["06_End_Market_Model"]
    data["end_units_full"] = read_table(ws, "Tank demand by end market (2025-2035, Units)")
    data["end_revenue_full"] = read_table(ws, "Revenue by end market (2025-2035, $Mn)")
    data["segment_growth"] = read_table(ws, "Segment growth summary (2025-2035, % CAGR)")
    data["segment_score"] = read_table(ws, "Segment profitability / complexity scorecard (base case)")
    data["named_count_end"] = read_table(ws, "Cumulative named project count by end market (2025-2035, count)")

    ws = wb["07_Regional_Model"]
    data["region_revenue_full"] = read_table(ws, "Market size by region (2025-2035, $Mn)")
    data["region_units_full"] = read_table(ws, "Tank demand by region (2025-2035, Units)")
    data["region_share_full"] = read_table(ws, "Regional market share (2025 vs 2035, % of revenue)")
    data["country_2035"] = read_table(ws, "Top 15 countries - indicative 2035 market opportunity (base case, $Mn)")

    ws = wb["08_Competitive_Landscape"]
    data["share_2025"] = read_table(ws, "Modeled market share by player (2025, % of global revenue)")
    data["share_2035"] = read_table(ws, "Modeled market share by player (2035, % of global revenue)")
    data["player_revenue"] = read_table(ws, "Modeled player revenue exposure (2025 / 2030 / 2035, $Mn)")
    data["capability"] = read_table(ws, "Capability matrix (1 = low, 5 = high)")

    ws = wb["10_Thyssenkrupp_Opportunity"]
    data["tk_addressable"] = read_table(ws, "thyssenkrupp addressable market (2025-2035, $Mn)")
    data["tk_end"] = read_table(ws, "Addressable market by end market (2025 / 2030 / 2035, $Mn)")
    data["tk_region"] = read_table(ws, "Addressable market by region (2025 / 2030 / 2035, $Mn)")
    data["tk_entry"] = read_table(ws, "thyssenkrupp revenue opportunity by entry path (2025-2035, $Mn)")
    data["tk_penetration"] = read_table(ws, "Penetration assumptions by entry path (2025 vs 2035, % of addressable market)")
    data["tk_gap"] = read_table(ws, "Capability gap scorecard (1 = weak, 5 = strong)")
    data["tk_partner"] = read_table(ws, "Potential partner archetypes for a partner-led route")

    return data

def find_workbook() -> Path | None:
    candidates = [Path("ammonia_storage_tank_market_model.xlsx"), Path(__file__).with_name("ammonia_storage_tank_market_model.xlsx"), Path.cwd() / "ammonia_storage_tank_market_model.xlsx"]
    for c in candidates:
        if c.exists(): return c
    return None

# =========================
# Render Views
# =========================
def render_overview(data, mult_factor):
    with st.spinner("Aggregating executive view..."):
        page_header("Executive Overview", "A curated view of market size, growth direction, scenario spread and structural mix across the ammonia storage tank opportunity.")

        rev = data["market_revenue"].iloc[0]
        units = data["tank_demand"].iloc[0]
        storage = data["storage_capacity"].iloc[0]
        tk_addr = data["tk_addr_cp"].iloc[0]

        c1, c2, c3, c4 = st.columns(4)
        with c1: card_metric("2025 Market Revenue", fmt_mn(float(rev["2025"])), "Preview view of starting market scale.")
        with c2: card_metric("2035 Market Revenue", fmt_mn(float(rev["2035"]), mult_factor), "Directional ceiling under current assumptions.")
        with c3: card_metric("2035 Tank Demand", fmt_units(float(units["2035"]), mult_factor), "Annual tank equivalents, rounded.")
        with c4: card_metric("2035 thyssenkrupp Addressable", fmt_mn(float(tk_addr["2035"]), mult_factor), "Total addressable revenue pool.")

        insight_text = "<strong>Strategic View:</strong> Under aggressive adoption parameters, market value accelerates heavily toward newbuild cracking and green-import infrastructure, dwarfing traditional fertilizer base-load." if mult_factor > 1.05 else "<strong>Strategic View:</strong> The market remains anchored by fertilizer replacement and brownfield activity, while green ammonia and ammonia cracking add visible long-cycle growth."
        st.markdown(f'<div class="insight-box">{insight_text}</div>', unsafe_allow_html=True)

        col1, col2 = st.columns([1.6, 1])
        with col1:
            section_open("Global market trajectory", "Revenue and tank-demand trend lines dynamically adjusted.")
            yrs = [int(c) for c in rev.index if c != "Metric"]
            rev_vals = [round(float(rev[str(y)]) * (mult_factor if y > 2025 else 1.0), 1) for y in yrs]
            unit_vals = [round(float(units[str(y)]) * (mult_factor if y > 2025 else 1.0), 1) for y in yrs]
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=yrs, y=rev_vals, mode="lines+markers", name="Market revenue ($Mn)", line=dict(color=BURGUNDY, width=3), marker=dict(size=8, color=BURGUNDY)))
            fig.add_trace(go.Bar(x=yrs, y=unit_vals, name="Tank demand (units)", marker_color=GOLD, opacity=0.6, yaxis="y2"))
            fig.update_layout(yaxis=dict(title="$Mn"), yaxis2=dict(title="Units", overlaying="y", side="right", showgrid=False), bargap=0.55)
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col2:
            section_open("Revenue mix shift", "End-market share comparison.")
            mix = data["end_share"].copy()
            mix_long = mix.melt(id_vars=get_col(mix, ["End market"]), value_vars=["2025", "2035"], var_name="Year", value_name="Share")
            fig = px.bar(mix_long, x="Year", y="Share", color=get_col(mix, ["End market"]), color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT, BURGUNDY_MID], barmode="stack")
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()
        page_footer()

def render_market_growth(data, mult_factor):
    with st.spinner("Calculating market trajectories..."):
        page_header("Market Size & Growth", "Preview of revenue, units, storage and pricing structure.")

        revenue = data["market_revenue"].iloc[0]
        units = data["tank_demand"].iloc[0]
        storage = data["storage_capacity"].iloc[0]

        selected = pd.DataFrame({
            "Metric": ["Market revenue ($Mn)", "Tank demand (Units)", "Storage capacity ('000 m³)"],
            "2025": [revenue["2025"], units["2025"], storage["2025"]],
            "2030": [float(revenue["2030"]) * (1+(mult_factor-1)/2), float(units["2030"])* (1+(mult_factor-1)/2), float(storage["2030"])* (1+(mult_factor-1)/2)],
            "2035": [float(revenue["2035"]) * mult_factor, float(units["2035"]) * mult_factor, float(storage["2035"]) * mult_factor],
        })
        for col in ["2025", "2030", "2035"]: selected[col] = pd.to_numeric(selected[col], errors="coerce").apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "—")

        col1, col2 = st.columns([1.35, 1])
        with col1:
            section_open("Selected-year headline snapshot", "Milestone years scaled by scenario logic.")
            st.dataframe(selected, use_container_width=True, hide_index=True)
            section_close()

        with col2:
            section_open("Base-case growth summary", "CAGR view derived from base calculation layer.")
            gs = data["growth_summary"].copy()
            metric_col = gs.columns[0]
            cagr_col = gs.columns[-1]
            for col in gs.columns:
                if col == metric_col: continue 
                elif col == cagr_col: gs[col] = (pd.to_numeric(gs[col], errors="coerce") * 100).round(1).astype(str) + "%"
                else: gs[col] = pd.to_numeric(gs[col], errors="coerce").apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "—")
            st.dataframe(gs, use_container_width=True, hide_index=True)
            section_close()

        col3, col4 = st.columns(2)
        with col3:
            section_open("Revenue by end market", "Preview of distribution.")
            rev_end = data["revenue_end"].copy()
            rev_end["2035"] = pd.to_numeric(rev_end["2035"], errors="coerce") * mult_factor
            fig = px.bar(rev_end, x=get_col(rev_end, ["End market"]), y=["2025", "2030", "2035"], barmode="group", color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT])
            fig.update_layout(xaxis_title="", yaxis_title="$Mn")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col4:
            section_open("Average tank cost by end market", "Rounded preview benchmark.")
            atc = data["avg_tank_cost"].copy()
            for col in ["2025", "2035"]:
                if col in atc.columns: atc[col] = pd.to_numeric(atc[col], errors="coerce").apply(lambda x: f"${x:,.1f}M" if pd.notna(x) else "—")
            st.dataframe(atc, use_container_width=True, hide_index=True)
            section_close()
        page_footer()

def render_pipeline(data):
    with st.spinner("Loading project database..."):
        page_header("Demand Drivers & Pipeline", "Project-led preview of where storage demand can emerge, using a curated subset of named projects.")
        project_db = data["project_db"].copy()
        
        col1, col2 = st.columns([1.25, 1])
        with col1:
            section_open("Named project timeline", "Bubble size reflects expected tank demand; color reflects end market.")
            fig = px.scatter(project_db, x=get_col(project_db, ["Start year"]), y=get_col(project_db, ["Capacity (Mtpa)"]), size=get_col(project_db, ["Expected tanks"]), color=get_col(project_db, ["End market"]), hover_name=get_col(project_db, ["Project"]), color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT, BURGUNDY_MID], size_max=36)
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col2:
            section_open("Pipeline by end market", "Risk-adjusted named capacity composition.")
            end = data["pipeline_end"].copy()
            cap_col = get_col(end, ["Risk-adjusted capacity (Mtpa)", "Risk adjusted capacity (Mtpa)", "Capacity (Mtpa)"])
            fig = px.pie(end, values=cap_col, names=get_col(end, ["End market"]), hole=0.6, color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT, BURGUNDY_MID])
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        st.markdown('<div class="insight-box"><strong>Pipeline Note:</strong> The full analytical environment contains broader modeled demand than the named-project list shown here.</div>', unsafe_allow_html=True)
        page_footer()

def render_regions(data, mult_factor):
    with st.spinner("Rendering geospatial data..."):
        page_header("Regional Analysis", "Preview of revenue concentration and country hotspots across the dashboard coverage universe.")

        region_rev = data["region_revenue_full"].copy()
        country = data["country_2035"].copy()

        section_open("Global Opportunity Heatmap (2035)", "Intensity of 2035 tank market opportunity by nation.")
        opportunity_col = get_col(country, ["2035 opportunity", "2035 market opportunity"])
        
        country["Scaled Opportunity"] = pd.to_numeric(country[opportunity_col], errors="coerce") * mult_factor
        
        fig = px.choropleth(
            country, 
            locations=get_col(country, ["Country"]), 
            locationmode="country names",
            color="Scaled Opportunity", 
            color_continuous_scale=[[0, "#FAF5F7"], [0.5, GOLD], [1, BURGUNDY]],
            projection="natural earth",
        )
        fig.update_layout(
            margin=dict(l=0, r=0, t=10, b=0),
            geo=dict(showframe=False, showcoastlines=True, coastlinecolor="rgba(0,0,0,0.08)", projection_scale=1.1, center=dict(lat=20, lon=0), bgcolor='rgba(0,0,0,0)')
        )
        fig.update_traces(hovertemplate="<b>%{location}</b><br>Opportunity: ≈$%{z:,.0f}M<extra></extra>", marker_line_color='rgba(255,255,255,0.5)', marker_line_width=0.5)
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        section_close()

        col1, col2 = st.columns([1.4, 1])
        with col1:
            section_open("Regional revenue trajectory", "Full trend line dynamically scaled.")
            id_col = get_col(region_rev, ["Region"])
            long_df = region_rev.melt(id_vars=id_col, var_name="Year", value_name="Revenue")
            long_df["Year"] = long_df["Year"].astype(int)
            long_df["Revenue"] = long_df.apply(lambda row: float(row["Revenue"]) * (mult_factor if row["Year"]>2025 else 1.0), axis=1)
            
            fig = px.line(long_df, x="Year", y="Revenue", color=id_col, color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT, BURGUNDY_MID, "#C5A0B6"], line_shape="spline")
            fig.update_traces(line=dict(width=3))
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()
            
        with col2:
            section_open("Top 2035 country hotspots", "Indicative ranking.")
            fig = px.bar(country.sort_values("Scaled Opportunity").tail(10), x="Scaled Opportunity", y=get_col(country, ["Country"]), orientation="h", color="Scaled Opportunity", color_continuous_scale=[[0, GOLD], [1, BURGUNDY]])
            fig.update_layout(coloraxis_showscale=False, xaxis_title="$Mn", yaxis_title="")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        insight_text = "<strong>Regional Edge:</strong> European and Asian import-terminal infrastructure heavily out-scales domestic US EPC logic under current aggressive scaling assumptions." if mult_factor > 1.05 else "<strong>Regional Takeaway:</strong> Europe and Japan matter disproportionately for cracking adjacencies, while North America and the Gulf drive pure EPC volume."
        st.markdown(f'<div class="insight-box">{insight_text}</div>', unsafe_allow_html=True)
        page_footer()

def render_segments(data):
    page_header("End-Market Segmentation", "Preview of fertilizer, green ammonia, bunkering and ammonia cracking demand without exposing the full year-by-year detail.")

    end_rev = data["end_revenue_full"].copy()
    end_units = data["end_units_full"].copy()
    score = data["segment_score"].copy()
    named_count = data["named_count_end"].copy()

    tab1, tab2, tab3 = st.tabs(["Revenue View", "Demand View", "Segment Scorecard"])

    with tab1:
        col1, col2 = st.columns([1.35, 1])
        with col1:
            section_open("Revenue by end market", "Full year trend is shown as a visual only.")
            id_col = get_col(end_rev, ["End market"])
            long_df = end_rev.melt(id_vars=id_col, var_name="Year", value_name="Revenue")
            long_df["Year"] = long_df["Year"].astype(int)
            fig = px.area(long_df, x="Year", y="Revenue", color=id_col, color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT, BURGUNDY_MID])
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()
        with col2:
            section_open("Selected-year revenue preview", "Rounded preview table for 2025 / 2030 / 2035.")
            sel = end_rev[[id_col, "2025", "2030", "2035"]].copy()
            for col in ["2025", "2030", "2035"]:
                sel[col] = pd.to_numeric(sel[col], errors="coerce").apply(lambda x: f"${x:,.0f}M" if pd.notna(x) else "—")
            st.dataframe(sel, use_container_width=True, hide_index=True)
            section_close()

    with tab2:
        col1, col2 = st.columns([1.2, 1])
        with col1:
            section_open("Tank demand by end market", "Physical units shown as rounded equivalents.")
            id_col = get_col(end_units, ["End market"])
            long_df = end_units.melt(id_vars=id_col, var_name="Year", value_name="Units")
            long_df["Year"] = long_df["Year"].astype(int)
            fig = px.bar(long_df[long_df["Year"].isin([2025, 2030, 2035])], x=id_col, y="Units", color="Year", barmode="group", color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT])
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()
        with col2:
            section_open("Named project count by end market", "Named-project count is used here as a signal of visible market momentum.")
            count_col = get_col(named_count, ["Project count", "Count"])
            fig = px.bar(named_count, x=get_col(named_count, ["End market"]), y=count_col, color=get_col(named_count, ["End market"]), color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_SOFT, BURGUNDY_MID])
            fig.update_layout(showlegend=False, xaxis_title="")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

    with tab3:
        col1, col2 = st.columns([1, 1.2])
        with col1:
            section_open("Complexity / profitability scorecard", "Qualitative preview view.")
            st.dataframe(score, use_container_width=True, hide_index=True)
            section_close()
        with col2:
            section_open("Segment growth summary", "CAGR is shown as directional segment ranking.")
            sg = data["segment_growth"].copy()
            dem_cagr = get_col(sg, ["Tank demand CAGR", "Demand CAGR"])
            sg["Display CAGR (%)"] = (pd.to_numeric(sg[dem_cagr], errors="coerce") * 100).round(1)
            fig = px.bar(sg.sort_values("Display CAGR (%)"), x="Display CAGR (%)", y=get_col(sg, ["End market"]), orientation="h", color="Display CAGR (%)", color_continuous_scale=[[0, GOLD], [1, BURGUNDY]])
            fig.update_layout(coloraxis_showscale=False, xaxis_title="CAGR (%)", yaxis_title="")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

    page_footer()

def render_competition(data):
    with st.spinner("Anonymizing competitor data..."):
        page_header("Competitive Landscape", "Preview of likely market control, player exposure and capability positioning across storage-led EPC scope. Competitor names have been anonymized for this preview.")

        s25 = data["share_2025"].copy()
        s35 = data["share_2035"].copy()
        pr = data["player_revenue"].copy()
        cap = data["capability"].copy()

        # --- ANONYMIZATION LOGIC ---
        player_col = get_col(s25, ["Player"])
        unique_players = s25[player_col].dropna().unique()
        
        player_map = {}
        comp_count = 1
        for p in unique_players:
            p_str = str(p).strip().lower()
            if p_str == "others":
                player_map[p_str] = "Others"
            elif "thyssenkrupp" in p_str:
                player_map[p_str] = "thyssenkrupp"
            else:
                player_map[p_str] = f"Company {comp_count}"
                comp_count += 1
                
        def apply_anon(val):
            key = str(val).strip().lower()
            return player_map.get(key, val)

        s25[player_col] = s25[player_col].apply(apply_anon)
        
        player_col_s35 = get_col(s35, ["Player"])
        s35[player_col_s35] = s35[player_col_s35].apply(apply_anon)
        
        player_col_pr = get_col(pr, ["Player"])
        pr[player_col_pr] = pr[player_col_pr].apply(apply_anon)
        
        player_col_cap = get_col(cap, ["Player"])
        cap[player_col_cap] = cap[player_col_cap].apply(apply_anon)
        # ---------------------------

        col1, col2 = st.columns([1, 1])
        with col1:
            section_open("2025 modeled market share", "Preview-level share view; not an audited historical market-share statement.")
            fig = px.pie(
                s25[s25[player_col] != "Others"],
                values=get_col(s25, ["2025 share"]), names=player_col, hole=0.6,
                color_discrete_sequence=[BURGUNDY, BURGUNDY_MID, GOLD, BURGUNDY_SOFT, "#D37A9E", "#C6A6B6", "#A68292"]
            )
            fig.update_traces(textinfo="percent")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col2:
            section_open("Share migration (2025 vs 2035)", "Side-by-side view of how the field may evolve.")
            m = s25.merge(s35, on=player_col, suffixes=("_2025", "_2035"))
            col25 = get_col(m, ["2025 share"])
            col35 = get_col(m, ["2035 share"])
            m = m[m[player_col] != "Others"].copy()
            
            m_long = m.melt(id_vars=[player_col], value_vars=[col25, col35], var_name="Year", value_name="Share")
            m_long["Year"] = m_long["Year"].apply(lambda x: "2025" if "2025" in str(x) else "2035")
            m_long["Share"] = (pd.to_numeric(m_long["Share"], errors="coerce") * 100).round(1)
            
            m_sorted = m.sort_values(col25, ascending=False)[player_col].tolist()
            
            fig = px.bar(
                m_long, 
                x=player_col, 
                y="Share", 
                color="Year", 
                barmode="group",
                color_discrete_sequence=[BURGUNDY, GOLD],
                category_orders={player_col: m_sorted}
            )
            fig.update_layout(xaxis_title="", yaxis_title="Share (%)", legend_title="")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        col3, col4 = st.columns([1.1, 1])
        with col3:
            section_open("Capability matrix", "Scaled 1–5 to show relative strength only.")
            heat = cap.set_index(player_col_cap)
            fig = px.imshow(
                heat.values, x=heat.columns, y=heat.index, aspect="auto",
                color_continuous_scale=[[0, "#FAF5F7"], [0.5, GOLD], [1, BURGUNDY]]
            )
            fig.update_layout(coloraxis_showscale=False)
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col4:
            section_open("Player revenue exposure", "Selected-year exposure from the modeled market view.")
            display = pr[pr[player_col_pr] != "Others"].copy()
            for col in ["2025", "2030", "2035"]:
                if col in display.columns:
                    display[col] = pd.to_numeric(display[col], errors="coerce").apply(lambda x: f"${x:,.0f}M" if pd.notna(x) else "—")
            st.dataframe(display, use_container_width=True, hide_index=True)
            section_close()

        st.markdown(
            '<div class="insight-box"><strong>Competition Summary:</strong> A select group of pure-play EPCs currently forms the core storage-heavy set, '
            'while thyssenkrupp is best positioned as an adjacent integrator leveraging its ammonia technology and cracking expertise, rather than competing day-one as a pure tank EPC leader.</div>',
            unsafe_allow_html=True,
        )
        page_footer()

def render_tk(data, mult_factor):
    with st.spinner("Extracting thyssenkrupp strategic views..."):
        page_header("thyssenkrupp Opportunity", "thyssenkrupp-specific preview of addressability, entry-path economics, and regional focus.")

        tk_addr = data["tk_addressable"].iloc[0]
        tk_entry = data["tk_entry"].copy()
        tk_end = data["tk_end"].copy()
        tk_region = data["tk_region"].copy()
        tk_gap = data["tk_gap"].copy()
        tk_pen = data["tk_penetration"].copy()
        tk_partner = data["tk_partner"].copy()

        c1, c2, c3, c4 = st.columns(4)
        with c1: card_metric("2025 Addressable Pool", fmt_mn(float(tk_addr["2025"])), "Addressable, not winnable.")
        with c2: card_metric("2030 Addressable Pool", fmt_mn(float(tk_addr["2030"]), mult_factor), "Preview milestone year.")
        with c3: card_metric("2035 Addressable Pool", fmt_mn(float(tk_addr["2035"]), mult_factor), "Scaled by custom parameters.")
        with c4: card_metric("Best Entry Route", "Partner-led", "Leverages process & tech strength.")

        col1, col2 = st.columns([1.35, 1])
        with col1:
            section_open("Revenue opportunity by entry path", "Entry paths scaling aggressively based on assumptions.")
            long_df = tk_entry.melt(id_vars=get_col(tk_entry, ["Year"]), value_vars=["Build", "Partner", "Acquire"], var_name="Entry path", value_name="Revenue")
            long_df["Revenue"] = pd.to_numeric(long_df["Revenue"], errors="coerce") * mult_factor
            fig = px.line(long_df, x=get_col(tk_entry, ["Year"]), y="Revenue", color="Entry path", color_discrete_map={"Build": BURGUNDY_SOFT, "Partner": BURGUNDY, "Acquire": GOLD}, line_shape="spline")
            fig.update_traces(line=dict(width=3))
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()
            
        with col2:
            section_open("Penetration assumptions", "Preview percentages only.")
            tk_pen = tk_pen.rename(columns={"2025": "2025 (%)", "2035": "2035 (%)"})
            tk_pen["2025 (%)"] = (pd.to_numeric(tk_pen["2025 (%)"], errors="coerce") * 100).round(1).astype(str) + "%"
            tk_pen["2035 (%)"] = (pd.to_numeric(tk_pen["2035 (%)"], errors="coerce") * 100).round(1).astype(str) + "%"
            st.dataframe(tk_pen, use_container_width=True, hide_index=True)
            section_close()

        col3, col4 = st.columns([1, 1])
        with col3:
            section_open("2035 addressable market by end market", "Best viewed as strategic focus ranking.")
            tk_end["2035"] = pd.to_numeric(tk_end["2035"], errors="coerce") * mult_factor
            fig = px.bar(tk_end.sort_values("2035"), x="2035", y=get_col(tk_end, ["End market"]), orientation="h", color="2035", color_continuous_scale=[[0, GOLD], [1, BURGUNDY]])
            fig.update_layout(coloraxis_showscale=False, xaxis_title="$Mn", yaxis_title="")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col4:
            section_open("2035 addressable market by region", "Regional prioritization lens for thyssenkrupp strategy.")
            tk_region["2035"] = pd.to_numeric(tk_region["2035"], errors="coerce") * mult_factor
            fig = px.bar(tk_region.sort_values("2035"), x="2035", y=get_col(tk_region, ["Region"]), orientation="h", color="2035", color_continuous_scale=[[0, GOLD], [1, BURGUNDY]])
            fig.update_layout(coloraxis_showscale=False, xaxis_title="$Mn", yaxis_title="")
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        col5, col6 = st.columns([1, 1.1])
        with col5:
            section_open("Capability gap scorecard", "Relative strength view: 1 weak / 5 strong.")
            fig = px.bar_polar(
                tk_gap,
                r=get_col(tk_gap, ["Score"]),
                theta=get_col(tk_gap, ["Capability"]),
                color=get_col(tk_gap, ["Capability"]),
                color_discrete_sequence=[BURGUNDY, GOLD, BURGUNDY_MID, BURGUNDY_SOFT, "#D7A2BA", "#C4A8B7"],
            )
            fig.update_layout(showlegend=False)
            st.plotly_chart(chart_theme(fig), use_container_width=True, config={"displayModeBar": False})
            section_close()

        with col6:
            section_open("Partner-led route archetypes", "Controlled preview of where partnerships could matter most.")
            st.dataframe(tk_partner, use_container_width=True, hide_index=True)
            section_close()

        insight_text = "<strong>thyssenkrupp Strategic Update:</strong> High-adoption parameters significantly widen the gap between Partner-led and Build-it-yourself strategies. Time-to-market is the primary risk factor." if mult_factor > 1.05 else "<strong>thyssenkrupp Strategic Update:</strong> The data reinforces expanding from process, green-ammonia, and cracking strengths into storage-led EPC through integration first."
        st.markdown(f'<div class="insight-box">{insight_text}</div>', unsafe_allow_html=True)
        page_footer()

def render_methodology(data):
    page_header(
        "Methodology & Coverage",
        "A concise explanation of how the dashboard analytics are built without exposing internal file structures, raw model architecture or deployment details.",
    )

    col1, col2 = st.columns(2)
    with col1:
        section_open("Modeling approach", "How the market view is constructed.")
        st.markdown(
            """
            **Bottom-up, project-linked demand framework**

            The market view is built by combining:
            - ammonia capacity expansion signals,
            - storage-day requirements by end market,
            - conversion of ammonia throughput into storage infrastructure,
            - tank sizing and configuration logic,
            - pricing benchmarks and cost curves.

            This approach is designed to reflect how storage demand emerges in practice rather than relying on a single top-down market extrapolation.
            """
        )
        section_close()

    with col2:
        section_open("Demand construction logic", "Where demand comes from.")
        st.markdown(
            """
            Demand is structured across four principal end markets:

            - **Fertilizer** — base-load storage and replacement demand  
            - **Green ammonia** — export/import and low-carbon infrastructure growth  
            - **Maritime bunkering** — emerging fuel-infrastructure demand  
            - **Ammonia cracking** — hydrogen-import and downstream conversion demand  

            Each segment is built using pipeline activity, risk-adjusted probability logic, storage intensity assumptions and segment-specific tank configuration rules.
            """
        )
        section_close()

    col3, col4 = st.columns(2)
    with col3:
        section_open("Project pipeline integration", "Why the dashboard can show direction without exposing the full underlying detail.")
        st.markdown(
            """
            A structured project database underpins visibility into market timing and location.

            The framework combines:
            - named ammonia production and terminal projects,
            - stage-based probability adjustment,
            - regional and segment allocation logic,
            - a secondary demand layer to account for smaller or less-visible opportunities.

            This allows the dashboard to show clear directional insight while still protecting granular underlying coverage.
            """
        )
        section_close()

    with col4:
        section_open("Pricing & revenue logic", "How physical demand becomes market value.")
        st.markdown(
            """
            Revenue is derived using:
            - average tank size assumptions,
            - technology-specific capex intensity,
            - end-market complexity modifiers,
            - regional price differentials,
            - inflation-linked pricing progression.

            This converts physical storage demand into revenue opportunity while preserving end-market and regional differentiation.
            """
        )
        section_close()

    col5, col6 = st.columns(2)
    with col5:
        section_open("Scenario framework", "How uncertainty is handled.")
        st.markdown(
            """
            Three scenario layers are used:

            - **Base case** — realistic project execution and moderate adoption  
            - **Upside** — faster green ammonia, import-terminal and bunkering execution  
            - **Downside** — project delays, slower adoption and tighter capex cycles  

            Scenario shifts influence tank demand, market size and regional mix.
            """
        )
        section_close()

    with col6:
        section_open("Preview data controls", "Why not all figures are shown here.")
        st.markdown(
            """
            This dashboard represents a **curated preview layer**.

            Displayed data is intentionally:
            - aggregated,
            - rounded,
            - filtered to selected milestone views,
            - limited in project and country granularity.

            The purpose is to show analytical depth and market structure without opening the full proprietary dataset.
            """
        )
        section_close()

    st.markdown(
        '<div class="insight-box"><strong>Strategic Market Research | Proprietary Market Methodology</strong><br>'
        'This preview environment is designed to communicate the logic, quality and structure of the underlying analysis while intentionally withholding the full depth of the detailed deliverables.</div>',
        unsafe_allow_html=True,
    )
    page_footer()

# =========================
# Main app logic & Security Gate
# =========================
def check_access():
    # Force to string and strip invisible spaces just in case it pulls weirdly from Streamlit secrets
    expected_password = str(st.secrets.get("ACCESS_CODE", "SMR2026")).strip()
    
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    brand_sidebar()

    with st.sidebar:
        st.markdown("<h3 style='color:var(--ink); font-weight:700; margin-bottom:12px;'>🔐 Access Login</h3>", unsafe_allow_html=True)
        
        # 1. Wrap in a form to prevent accidental reruns and capture keyboard "Enter" hits safely
        with st.form("login_form"):
            name = st.text_input("Name*")
            email = st.text_input("Email*")
            company = st.text_input("Company / Organization")
            password = st.text_input("Access Code*", type="password")
            st.markdown("<br>", unsafe_allow_html=True)
            enter = st.form_submit_button("Secure Login", use_container_width=True)

    if enter:
        # 2. Strip whitespace from all user inputs to prevent trailing-space errors
        clean_name = name.strip()
        clean_email = email.strip()
        clean_pass = password.strip()
        
        if not clean_name or not clean_email or not clean_pass:
            st.sidebar.warning("⚠️ Please fill in your Name, Email, and Access Code.")
        elif clean_pass != expected_password:
            st.sidebar.error("❌ Invalid Access Code. Please try again.")
        else:
            st.session_state.authenticated = True
            st.session_state.viewer_name = clean_name
            st.session_state.viewer_company = company.strip()
            log_access(clean_name, clean_email)
            st.rerun()

    if not st.session_state.authenticated:
        st.markdown(
            """
            <div style="display:flex; flex-direction:column; align-items:center; justify-content:center; height:60vh;">
                <h2 style='color:var(--burgundy); font-weight:800; font-size:2.4rem;'>Dashboard Secured</h2>
                <p style='color:var(--muted); font-size:1.1rem;'>Please use the sidebar to authenticate and load the market view.</p>
            </div>
            """, unsafe_allow_html=True)
        st.stop()
    return True

check_access()

# =========================
# Authenticated Shell
# =========================
workbook_path = find_workbook()
if workbook_path is None:
    st.error("Preview data source not found. Please place the dashboard data file in the application folder.")
    st.stop()

data = load_model(str(workbook_path))
brand_sidebar()

# Identify User
viewer_name = st.session_state.get("viewer_name", "Guest")
viewer_company = st.session_state.get("viewer_company", "")
viewer_chip = viewer_name + (f" | {viewer_company}" if viewer_company else "")
st.sidebar.markdown(f'<div class="viewer-chip">Verified: {viewer_chip}</div>', unsafe_allow_html=True)

# ⚙️ INTERACTIVE WHAT-IF ENGINE
st.sidebar.markdown("<br>", unsafe_allow_html=True)
st.sidebar.markdown("<p style='font-weight:700; color:var(--burgundy); margin-bottom:4px;'>⚙️ Scenario Engine</p>", unsafe_allow_html=True)
slider_val = st.sidebar.slider("Green Ammonia Adoption Shift", min_value=-20, max_value=30, value=0, step=5, format="%d%%")
engine_multiplier = 1.0 + (slider_val / 100.0)

page = st.sidebar.radio(
    "",
    [
        "Executive Overview",
        "Market Size & Growth",
        "Demand Drivers & Pipeline",
        "End-Market Segmentation",
        "Regional Analysis",
        "Competitive Landscape",
        "thyssenkrupp Opportunity",
        "Methodology & Coverage"
    ],
)

st.sidebar.markdown("---")
if st.sidebar.button("End Session", use_container_width=True):
    st.session_state.authenticated = False
    st.rerun()

# Router
if page == "Executive Overview": render_overview(data, engine_multiplier)
elif page == "Market Size & Growth": render_market_growth(data, engine_multiplier)
elif page == "Demand Drivers & Pipeline": render_pipeline(data)
elif page == "End-Market Segmentation": render_segments(data)
elif page == "Regional Analysis": render_regions(data, engine_multiplier)
elif page == "Competitive Landscape": render_competition(data)
elif page == "thyssenkrupp Opportunity": render_tk(data, engine_multiplier)
elif page == "Methodology & Coverage": render_methodology(data)
